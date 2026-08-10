# -*- coding: utf-8 -*-
"""Standalone integration sweep for the production wire-sizing engine.

This is intentionally an IDE-runnable test/report workflow.  It does not
reimplement conductor, ampacity, voltage-drop, breaker, or EGC rules.  Each
test point creates a synthetic circuit-shaped input and calls
``CEDElectrical.Model.CircuitBranch`` directly.

Run from the repository root, for example::

    py -3 CEDLib.lib/UnitTests/wire_sizing_sweep.py

The runner uses the production model's breaker table and conductor ladder at
runtime.  When an IDE Python environment has ``xlsxwriter`` or ``openpyxl``
installed it writes the requested three-sheet XLSX report.  Otherwise it
writes three Excel-readable CSV files with the same report sections.
"""

from __future__ import print_function

import argparse
import csv
import importlib
import logging
import os
import sys
import types

TEST_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.abspath(os.path.join(TEST_DIR, os.pardir, os.pardir))
CEDLIB_ROOT = os.path.join(REPO_ROOT, "CEDLib.lib")
DEFAULT_OUTPUT = os.path.join(TEST_DIR, "output", "wire_sizing_sweep.xlsx")


# These are the six requested electrical inputs.  Voltage and poles are fed
# to the production circuit properties; CircuitBranch derives its phase from
# poles exactly as it does for a Revit ElectricalSystem.
ELECTRICAL_CONFIGURATIONS = (
    {"label": "120 V / 1 pole", "voltage": 120.0, "poles": 1},
    {"label": "277 V / 1 pole", "voltage": 277.0, "poles": 1},
    {"label": "208 V / 2 pole", "voltage": 208.0, "poles": 2},
    {"label": "208 V / 3 pole", "voltage": 208.0, "poles": 3},
    {"label": "480 V / 2 pole", "voltage": 480.0, "poles": 2},
    {"label": "480 V / 3 pole", "voltage": 480.0, "poles": 3},
)

LOAD_PERCENTAGES = tuple(range(10, 101, 10))
COARSE_LENGTH_START_FT = 1
COARSE_LENGTH_STEP_FT = 25
REFINEMENT_STEP_FT = 1
MAX_SEARCH_LENGTH_FT = 10000

COARSE_HEADERS = (
    "Breaker Rating",
    "Load",
    "Length",
    "Hot Wire Size",
    "Ground Wire Size",
    "Poles",
    "Voltage",
    "Load Percent",
    "Baseline Hot Wire Size",
    "Hot Upsize Steps",
    "Test Point Type",
    "Status / Notes",
    "Configuration",
    "Hot Wire Quantity",
    "Ground Wire Quantity",
    "Parallel Sets",
    "Wire Material",
    "Wire Temperature",
    "Base Ampacity",
    "Voltage Drop %",
)

TRANSITION_HEADERS = COARSE_HEADERS + (
    "Transition From Length",
    "Transition To Length",
    "Transition From Hot Wire Size",
    "Transition To Hot Wire Size",
    "Refinement Step",
)

SUMMARY_HEADERS = (
    "Breaker Rating",
    "Voltage",
    "Poles",
    "Configuration",
    "Baseline Hot Wire Size",
    "Maximum Tested Length",
    "Maximum Hot Wire Size",
    "Hot Upsize Steps Reached",
    "Five Upsizes Reached",
    "Adaptive Search Status",
    "Status / Notes",
    "Coarse Row Count",
    "Refinement Row Count",
    "Failure Row Count",
)


def _install_ide_shims():
    """Install the smallest Revit/pyRevit surface needed by CircuitBranch.

    The production model is deliberately Revit-facing, but its calculation
    methods are deterministic once circuit inputs are present.  These shims
    let an IDE process supply those inputs without opening Revit.  No sizing
    behavior is copied here.
    """

    try:
        import Autodesk.Revit.DB.Electrical  # noqa: F401
        from pyrevit import DB, revit, script  # noqa: F401
        from System import Guid  # noqa: F401
        return
    except Exception:
        pass

    class _ValueAccessor(object):
        def __init__(self, attribute):
            self.attribute = attribute

        def __get__(self, instance, owner=None):
            if instance is None:
                return self
            return getattr(instance, self.attribute, None)

    class _CircuitType(object):
        Circuit = "Circuit"
        Spare = "Spare"
        Space = "Space"

    class _StorageType(object):
        String = "String"
        Integer = "Integer"
        Double = "Double"
        ElementId = "ElementId"

    class _UnitTypeId(object):
        Volts = "Volts"
        VoltAmperes = "VoltAmperes"

    class _UnitUtils(object):
        @staticmethod
        def ConvertFromInternalUnits(value, unit_type):
            return value

    class _BuiltInParameter(object):
        RBS_ELEC_VOLTAGE = "RBS_ELEC_VOLTAGE"
        RBS_ELEC_CIRCUIT_NOTES_PARAM = "RBS_ELEC_CIRCUIT_NOTES_PARAM"
        RBS_ELEC_PANEL_TOTALESTLOAD_PARAM = "RBS_ELEC_PANEL_TOTALESTLOAD_PARAM"
        RBS_ELEC_PANEL_TOTAL_DEMAND_CURRENT_PARAM = "RBS_ELEC_PANEL_TOTAL_DEMAND_CURRENT_PARAM"
        RBS_FAMILY_CONTENT_DISTRIBUTION_SYSTEM = "RBS_FAMILY_CONTENT_DISTRIBUTION_SYSTEM"
        FAMILY_CONTENT_PART_TYPE = "FAMILY_CONTENT_PART_TYPE"

    class _ElectricalSystem(object):
        ApparentCurrent = _ValueAccessor("_apparent_current")
        ApparentLoad = _ValueAccessor("_apparent_load")
        PolesNumber = _ValueAccessor("_poles")
        PowerFactor = _ValueAccessor("_power_factor")

    class _DistributionSysType(object):
        pass

    db_module = types.ModuleType("Autodesk.Revit.DB")
    db_module.StorageType = _StorageType
    db_module.UnitTypeId = _UnitTypeId
    db_module.UnitUtils = _UnitUtils
    db_module.BuiltInParameter = _BuiltInParameter
    db_module.FamilyInstance = type("FamilyInstance", (object,), {})

    dbe_module = types.ModuleType("Autodesk.Revit.DB.Electrical")
    dbe_module.CircuitType = _CircuitType
    dbe_module.ElectricalSystem = _ElectricalSystem
    dbe_module.DistributionSysType = _DistributionSysType

    revit_package = types.ModuleType("Autodesk.Revit")
    revit_package.DB = db_module
    db_module.Electrical = dbe_module
    autodesk_module = types.ModuleType("Autodesk")
    autodesk_module.Revit = revit_package

    sys.modules["Autodesk"] = autodesk_module
    sys.modules["Autodesk.Revit"] = revit_package
    sys.modules["Autodesk.Revit.DB"] = db_module
    sys.modules["Autodesk.Revit.DB.Electrical"] = dbe_module

    logger = logging.getLogger("wire_sizing_sweep")
    script_module = types.ModuleType("pyrevit.script")
    script_module.get_logger = lambda: logger
    script_module.get_output = lambda: None
    revit_module = types.ModuleType("pyrevit.revit")
    revit_module.doc = None
    pyrevit_module = types.ModuleType("pyrevit")
    pyrevit_module.DB = db_module
    pyrevit_module.script = script_module
    pyrevit_module.revit = revit_module
    sys.modules["pyrevit"] = pyrevit_module
    sys.modules["pyrevit.script"] = script_module
    sys.modules["pyrevit.revit"] = revit_module

    system_module = types.ModuleType("System")
    system_module.Guid = lambda value: value
    sys.modules["System"] = system_module

    snippets_module = types.ModuleType("Snippets")
    elecutils_module = types.ModuleType("Snippets._elecutils")
    design_options_module = types.ModuleType("Snippets.design_options")
    revit_helpers_module = types.ModuleType("Snippets.revit_helpers")

    elecutils_module.is_circuit_eligible = lambda circuit: bool(
        circuit is not None
        and getattr(circuit, "CircuitType", None) == _CircuitType.Circuit
    )
    design_options_module.is_main_model_element = lambda element: True
    revit_helpers_module.get_elementid_value = lambda value: int(value or 0)
    revit_helpers_module.elementid_from_value = lambda value: int(value or 0)

    snippets_module._elecutils = elecutils_module
    snippets_module.design_options = design_options_module
    snippets_module.revit_helpers = revit_helpers_module
    sys.modules["Snippets"] = snippets_module
    sys.modules["Snippets._elecutils"] = elecutils_module
    sys.modules["Snippets.design_options"] = design_options_module
    sys.modules["Snippets.revit_helpers"] = revit_helpers_module


def _load_production_api():
    """Return the production model and its runtime sizing configuration."""

    _install_ide_shims()
    if CEDLIB_ROOT not in sys.path:
        sys.path.insert(0, CEDLIB_ROOT)

    from CEDElectrical.Model.CircuitBranch import (  # pylint: disable=import-outside-toplevel
        ALLOWED_WIRE_SIZES,
        CircuitBranch,
    )
    from CEDElectrical.Model.circuit_settings import (  # pylint: disable=import-outside-toplevel
        CircuitSettings,
    )
    from CEDElectrical.refdata.standard_ocp_table import (  # pylint: disable=import-outside-toplevel
        BREAKER_FRAME_SWITCH_TABLE,
    )
    from Autodesk.Revit.DB import Electrical as DBE  # pylint: disable=import-outside-toplevel

    return {
        "CircuitBranch": CircuitBranch,
        "CircuitSettings": CircuitSettings,
        "DBE": DBE,
        "breakers": tuple(sorted(int(key) for key in BREAKER_FRAME_SWITCH_TABLE.keys())),
        "conductor_order": tuple(ALLOWED_WIRE_SIZES),
    }


class _SyntheticParameter(object):
    def __init__(self, value, storage_type):
        self.value = value
        self.StorageType = storage_type
        self.HasValue = value is not None

    def AsDouble(self):
        return float(self.value)

    def AsInteger(self):
        return int(self.value)

    def AsString(self):
        return "" if self.value is None else str(self.value)


class _SyntheticPanel(object):
    Name = "IDE Sweep Panel"


class _SyntheticCircuit(object):
    """Minimal ElectricalSystem-shaped input object for CircuitBranch."""

    _next_id = 1

    def __init__(self, production, breaker_rating, load_current, length_ft, voltage, poles):
        self.Id = _SyntheticCircuit._next_id
        _SyntheticCircuit._next_id += 1
        self.BaseEquipment = _SyntheticPanel()
        self.CircuitNumber = str(self.Id)
        self.CircuitType = production["DBE"].CircuitType.Circuit
        self.Elements = ()
        self.Frame = None
        self.Length = float(length_ft)
        self.LoadName = "IDE Wire Sizing Sweep"
        self.Rating = float(breaker_rating)
        self._apparent_current = float(load_current)
        self._apparent_load = 0.0
        self._poles = int(poles)
        self._power_factor = None
        self._voltage = float(voltage)
        self._db = sys.modules["Autodesk.Revit.DB"]

    def get_Parameter(self, parameter):
        if parameter == self._db.BuiltInParameter.RBS_ELEC_VOLTAGE:
            return _SyntheticParameter(self._voltage, self._db.StorageType.Double)
        return None

    def LookupParameter(self, name):
        return None


def _normalise_wire_size(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.startswith("#"):
        text = text[1:]
    return text.strip()


def _wire_index(value, conductor_order):
    normalized = _normalise_wire_size(value)
    try:
        return conductor_order.index(normalized)
    except ValueError:
        return None


def _wire_label(value):
    if value is None:
        return ""
    return str(value)


def _notice_text(branch):
    notices = []
    try:
        items = branch.collect_qc_notices()
    except Exception as error:
        return "Notice collection failed: {}".format(error)
    for item in list(items or []):
        try:
            severity, message = str(item[1] or "").upper(), str(item[3] or "")
            if message:
                notices.append("{}: {}".format(severity, message))
        except Exception:
            notices.append(str(item))
    return "; ".join(notices)


def _empty_output_values():
    return {
        "Hot Wire Size": "",
        "Ground Wire Size": "",
        "Hot Wire Quantity": None,
        "Ground Wire Quantity": None,
        "Parallel Sets": None,
        "Wire Material": "",
        "Wire Temperature": "",
        "Base Ampacity": None,
        "Voltage Drop %": None,
    }


def run_production_sizing(production, breaker_rating, load_current, length_ft, voltage, poles):
    """Call the production sizing pipeline once and capture its outputs."""

    values = _empty_output_values()
    status = "OK"
    notes = ""
    branch = None
    try:
        circuit = _SyntheticCircuit(
            production,
            breaker_rating,
            load_current,
            length_ft,
            voltage,
            poles,
        )
        branch = production["CircuitBranch"](
            circuit,
            settings=production["CircuitSettings"](),
            # Force the same automatic-sizing path used by calculate-circuits.
            preview_values={"CKT_User Override_CED": 0},
        )
        branch.calculate_hot_wire_size()
        branch.calculate_neutral_wire_size()
        branch.calculate_ground_wire_size()
        branch.calculate_isolated_ground_wire_size()

        values.update(
            {
                "Hot Wire Size": _wire_label(getattr(branch, "hot_wire_size", "")),
                "Ground Wire Size": _wire_label(getattr(branch, "ground_wire_size", "")),
                "Hot Wire Quantity": getattr(branch, "hot_wire_quantity", None),
                "Ground Wire Quantity": getattr(branch, "ground_wire_quantity", None),
                "Parallel Sets": getattr(branch, "number_of_sets", None),
                "Wire Material": _wire_label(getattr(branch, "wire_material", "")),
                "Wire Temperature": _wire_label(getattr(branch, "wire_temp_rating", "")),
                "Base Ampacity": getattr(branch, "circuit_base_ampacity", None),
                "Voltage Drop %": (
                    100.0 * float(getattr(branch, "voltage_drop_percentage", 0.0))
                    if getattr(branch, "voltage_drop_percentage", None) is not None
                    else None
                ),
            }
        )
        if bool(getattr(branch, "calc_failed", False)):
            status = "CALC_FAILED"
        notes = _notice_text(branch)
        if status == "CALC_FAILED" and not notes:
            notes = "Production sizing marked this point as failed."
    except Exception as error:  # Keep one bad point from hiding the sweep.
        status = "EXCEPTION"
        notes = "{}: {}".format(type(error).__name__, error)

    return {
        "values": values,
        "status": status,
        "notes": notes,
        "branch": branch,
    }


def _make_row(
    production,
    breaker_rating,
    load_percent,
    length_ft,
    config,
    baseline_hot,
    test_point_type,
    result,
):
    values = dict(result.get("values") or _empty_output_values())
    hot = values.get("Hot Wire Size") or ""
    steps = None
    baseline_index = _wire_index(baseline_hot, production["conductor_order"])
    hot_index = _wire_index(hot, production["conductor_order"])
    if baseline_index is not None and hot_index is not None:
        steps = hot_index - baseline_index

    notes = str(result.get("notes") or "")
    status = str(result.get("status") or "EXCEPTION")
    status_notes = status if not notes else "{} - {}".format(status, notes)

    return {
        "Breaker Rating": breaker_rating,
        "Load": float(breaker_rating) * float(load_percent) / 100.0,
        "Length": int(length_ft),
        "Hot Wire Size": hot,
        "Ground Wire Size": values.get("Ground Wire Size") or "",
        "Poles": config["poles"],
        "Voltage": config["voltage"],
        "Load Percent": load_percent,
        "Baseline Hot Wire Size": baseline_hot or "",
        "Hot Upsize Steps": steps,
        "Test Point Type": test_point_type,
        "Status / Notes": status_notes,
        "Configuration": config["label"],
        "Hot Wire Quantity": values.get("Hot Wire Quantity"),
        "Ground Wire Quantity": values.get("Ground Wire Quantity"),
        "Parallel Sets": values.get("Parallel Sets"),
        "Wire Material": values.get("Wire Material") or "",
        "Wire Temperature": values.get("Wire Temperature") or "",
        "Base Ampacity": values.get("Base Ampacity"),
        "Voltage Drop %": values.get("Voltage Drop %"),
        "_status": status,
        "_breaker": breaker_rating,
        "_config": config["label"],
        "_voltage": config["voltage"],
        "_poles": config["poles"],
        "_load_percent": load_percent,
    }


def _determine_max_length(production, breaker_rating, config):
    """Find the adaptive maximum at 100% breaker load."""

    baseline_result = run_production_sizing(
        production,
        breaker_rating,
        breaker_rating,
        COARSE_LENGTH_START_FT,
        config["voltage"],
        config["poles"],
    )
    baseline_hot = baseline_result["values"].get("Hot Wire Size") or ""
    baseline_index = _wire_index(baseline_hot, production["conductor_order"])
    five_upsize_target = (
        baseline_index + 5
        if baseline_index is not None
        and baseline_index + 5 < len(production["conductor_order"])
        else None
    )
    # If five larger ladder entries do not exist, still run the adaptive
    # search until the production ladder's actual upper conductor is reached.
    # This preserves the requested upper-limit condition instead of stopping
    # immediately just because the target is unreachable.
    target_index = (
        five_upsize_target
        if five_upsize_target is not None
        else len(production["conductor_order"]) - 1
    )

    if baseline_index is None:
        return {
            "baseline_hot": baseline_hot,
            "baseline_result": baseline_result,
            "max_length": COARSE_LENGTH_START_FT,
            "max_result": baseline_result,
            "steps": None,
            "reached_five": False,
            "search_status": "BASELINE_UNAVAILABLE",
            "search_notes": "Baseline did not return a conductor in the production ladder.",
        }

    max_result = baseline_result
    max_length = COARSE_LENGTH_START_FT
    last_steps = 0
    for length_ft in range(COARSE_LENGTH_STEP_FT, MAX_SEARCH_LENGTH_FT + 1, COARSE_LENGTH_STEP_FT):
        result = run_production_sizing(
            production,
            breaker_rating,
            breaker_rating,
            length_ft,
            config["voltage"],
            config["poles"],
        )
        max_result = result
        max_length = length_ft
        current_hot = result["values"].get("Hot Wire Size") or ""
        current_index = _wire_index(current_hot, production["conductor_order"])
        if current_index is not None:
            last_steps = current_index - baseline_index
            if current_index >= target_index:
                reached_five = five_upsize_target is not None
                return {
                    "baseline_hot": baseline_hot,
                    "baseline_result": baseline_result,
                    "max_length": max_length,
                    "max_result": max_result,
                    "steps": last_steps,
                    "reached_five": reached_five,
                    "search_status": (
                        "REACHED_FIVE_UPSIZES"
                        if reached_five
                        else "UPPER_LIMIT_FIRST"
                    ),
                    "search_notes": (
                        "Production conductor ladder ended before five additional sizes were available."
                        if not reached_five
                        else ""
                    ),
                }
        if result["status"] in ("CALC_FAILED", "EXCEPTION"):
            return {
                "baseline_hot": baseline_hot,
                "baseline_result": baseline_result,
                "max_length": max_length,
                "max_result": max_result,
                "steps": last_steps,
                "reached_five": False,
                "search_status": "UPPER_LIMIT_FIRST",
                "search_notes": "Production sizing stopped the adaptive search at this point: {}.".format(
                    result["status"]
                ),
            }

    return {
        "baseline_hot": baseline_hot,
        "baseline_result": baseline_result,
        "max_length": max_length,
        "max_result": max_result,
        "steps": last_steps,
        "reached_five": False,
        "search_status": "SAFETY_GUARD_REACHED",
        "search_notes": "Adaptive search reached MAX_SEARCH_LENGTH_FT={}.".format(
            MAX_SEARCH_LENGTH_FT
        ),
    }


def _coarse_lengths(max_length):
    lengths = [COARSE_LENGTH_START_FT]
    lengths.extend(
        range(COARSE_LENGTH_STEP_FT, int(max_length) + 1, COARSE_LENGTH_STEP_FT)
    )
    return lengths


def _refinement_rows_for_group(
    production,
    breaker_rating,
    config,
    baseline_hot,
    coarse_rows,
):
    rows = []
    by_load = {}
    for row in coarse_rows:
        by_load.setdefault(row["_load_percent"], []).append(row)

    for load_percent, load_rows in sorted(by_load.items()):
        ordered = sorted(load_rows, key=lambda item: item["Length"])
        for left, right in zip(ordered, ordered[1:]):
            left_hot = left.get("Hot Wire Size") or ""
            right_hot = right.get("Hot Wire Size") or ""
            if not left_hot or not right_hot or left_hot == right_hot:
                continue

            for length_ft in range(
                int(left["Length"]) + REFINEMENT_STEP_FT,
                int(right["Length"]),
                REFINEMENT_STEP_FT,
            ):
                result = run_production_sizing(
                    production,
                    breaker_rating,
                    float(breaker_rating) * float(load_percent) / 100.0,
                    length_ft,
                    config["voltage"],
                    config["poles"],
                )
                row = _make_row(
                    production,
                    breaker_rating,
                    load_percent,
                    length_ft,
                    config,
                    baseline_hot,
                    "Refinement",
                    result,
                )
                row.update(
                    {
                        "Transition From Length": left["Length"],
                        "Transition To Length": right["Length"],
                        "Transition From Hot Wire Size": left_hot,
                        "Transition To Hot Wire Size": right_hot,
                        "Refinement Step": REFINEMENT_STEP_FT,
                    }
                )
                rows.append(row)
    return rows


def generate_sweep(production=None, progress=True):
    """Generate coarse rows, refined transition rows, and summary rows."""

    production = production or _load_production_api()
    breakers = production["breakers"]
    coarse_rows = []
    transition_rows = []
    summary_rows = []
    total_combinations = len(breakers) * len(ELECTRICAL_CONFIGURATIONS)
    combination_number = 0

    for breaker_rating in breakers:
        for config in ELECTRICAL_CONFIGURATIONS:
            combination_number += 1
            if progress:
                print(
                    "Processing combination {}/{}: {} A, {}".format(
                        combination_number,
                        total_combinations,
                        breaker_rating,
                        config["label"],
                    )
                )

            adaptive = _determine_max_length(production, breaker_rating, config)
            baseline_hot = adaptive["baseline_hot"]
            combination_rows = []
            for load_percent in LOAD_PERCENTAGES:
                for length_ft in _coarse_lengths(adaptive["max_length"]):
                    load_current = float(breaker_rating) * float(load_percent) / 100.0
                    result = run_production_sizing(
                        production,
                        breaker_rating,
                        load_current,
                        length_ft,
                        config["voltage"],
                        config["poles"],
                    )
                    row = _make_row(
                        production,
                        breaker_rating,
                        load_percent,
                        length_ft,
                        config,
                        baseline_hot,
                        "Coarse",
                        result,
                    )
                    combination_rows.append(row)
                    coarse_rows.append(row)

            refined = _refinement_rows_for_group(
                production,
                breaker_rating,
                config,
                baseline_hot,
                combination_rows,
            )
            transition_rows.extend(refined)

            all_combination_rows = combination_rows + refined
            hot_indices = [
                _wire_index(row.get("Hot Wire Size"), production["conductor_order"])
                for row in all_combination_rows
                if _wire_index(row.get("Hot Wire Size"), production["conductor_order"]) is not None
            ]
            max_hot = (
                production["conductor_order"][max(hot_indices)]
                if hot_indices
                else ""
            )
            max_steps = (
                max(hot_indices) - _wire_index(baseline_hot, production["conductor_order"])
                if hot_indices and _wire_index(baseline_hot, production["conductor_order"]) is not None
                else adaptive["steps"]
            )
            failure_count = sum(
                1
                for row in all_combination_rows
                if row.get("_status") in ("CALC_FAILED", "EXCEPTION")
            )
            search_note = adaptive["search_notes"]
            if failure_count:
                failure_note = "{} generated row failure(s).".format(failure_count)
                search_note = "; ".join([item for item in (search_note, failure_note) if item])

            summary_rows.append(
                {
                    "Breaker Rating": breaker_rating,
                    "Voltage": config["voltage"],
                    "Poles": config["poles"],
                    "Configuration": config["label"],
                    "Baseline Hot Wire Size": baseline_hot,
                    "Maximum Tested Length": adaptive["max_length"],
                    "Maximum Hot Wire Size": max_hot,
                    "Hot Upsize Steps Reached": max_steps,
                    "Five Upsizes Reached": "Yes" if adaptive["reached_five"] else "No",
                    "Adaptive Search Status": adaptive["search_status"],
                    "Status / Notes": search_note,
                    "Coarse Row Count": len(combination_rows),
                    "Refinement Row Count": len(refined),
                    "Failure Row Count": failure_count,
                }
            )

    result = {
        "production": production,
        "coarse_rows": coarse_rows,
        "transition_rows": transition_rows,
        "summary_rows": summary_rows,
    }
    result["validation_failures"] = validate_result(result)
    return result


def _iter_data_rows(result):
    for row in result.get("coarse_rows", []):
        yield row
    for row in result.get("transition_rows", []):
        yield row


def validate_result(result):
    """Return structural integrity failures without asserting electrical answers."""

    failures = []
    production = result["production"]
    expected_breakers = set(production["breakers"])
    expected_configs = set(config["label"] for config in ELECTRICAL_CONFIGURATIONS)
    summary_breakers = set(row["Breaker Rating"] for row in result["summary_rows"])
    summary_configs = set(row["Configuration"] for row in result["summary_rows"])

    if summary_breakers != expected_breakers:
        failures.append("Supported breaker coverage mismatch.")
    if summary_configs != expected_configs:
        failures.append("Electrical configuration coverage mismatch.")
    if len(result["summary_rows"]) != len(expected_breakers) * len(expected_configs):
        failures.append("Summary combination count mismatch.")

    rows_by_group = {}
    for row in result["coarse_rows"]:
        key = (
            row["Breaker Rating"],
            row["Configuration"],
            row["Load Percent"],
        )
        rows_by_group.setdefault(key, []).append(row)
        for header in COARSE_HEADERS:
            if row.get(header) is None and header in (
                "Breaker Rating",
                "Load",
                "Length",
                "Poles",
                "Voltage",
                "Load Percent",
                "Test Point Type",
                "Status / Notes",
            ):
                failures.append("Required value missing in coarse row: {}.".format(header))
        if row.get("Test Point Type") != "Coarse":
            failures.append("Coarse row has the wrong test point type.")

    for key, rows in rows_by_group.items():
        lengths = [row["Length"] for row in rows]
        if lengths != sorted(lengths) or len(lengths) != len(set(lengths)):
            failures.append("Lengths are not strictly increasing for {}.".format(key))

        previous_index = None
        for row in sorted(rows, key=lambda item: item["Length"]):
            current_index = _wire_index(row.get("Hot Wire Size"), production["conductor_order"])
            if current_index is None:
                continue
            if previous_index is not None and current_index < previous_index:
                failures.append("Hot conductor became smaller as length increased for {}.".format(key))
                break
            previous_index = current_index

    for row in result["transition_rows"]:
        if row.get("Test Point Type") != "Refinement":
            failures.append("Transition row has the wrong test point type.")
        if not (
            row.get("Transition From Length") < row.get("Length") < row.get("Transition To Length")
        ):
            failures.append("Refinement point is outside its transition interval.")
        if row.get("Refinement Step") not in (1, 5):
            failures.append("Refinement step is not 1 ft or 5 ft.")

    for row in _iter_data_rows(result):
        if row.get("Hot Upsize Steps") is not None:
            baseline_index = _wire_index(row.get("Baseline Hot Wire Size"), production["conductor_order"])
            hot_index = _wire_index(row.get("Hot Wire Size"), production["conductor_order"])
            if baseline_index is None or hot_index is None:
                failures.append("Hot Upsize Steps exists without valid production ladder sizes.")
            elif row["Hot Upsize Steps"] != hot_index - baseline_index:
                failures.append("Hot Upsize Steps is inconsistent with the production ladder.")

    for summary in result["summary_rows"]:
        if summary["Maximum Tested Length"] > MAX_SEARCH_LENGTH_FT:
            failures.append("Adaptive search exceeded its safety guard.")
        if summary["Maximum Tested Length"] < COARSE_LENGTH_START_FT:
            failures.append("Adaptive search produced an invalid maximum length.")

    return sorted(set(failures))


def assert_result_integrity(result):
    """Assertion entry point for IDE test runners and strict command-line use."""

    failures = result.get("validation_failures") or validate_result(result)
    assert not failures, "Wire sizing sweep structural integrity failures: {}".format(
        " | ".join(failures)
    )


def _strip_private(row, headers):
    return [row.get(header) for header in headers]


def _write_csv(path, headers, rows):
    with open(path, "w", newline="", encoding="utf-8-sig") as stream:
        writer = csv.writer(stream)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(_strip_private(row, headers))


def _safe_sheet_name(value):
    return value[:31]


def _write_xlsx_xlsxwriter(path, result):
    import xlsxwriter  # pylint: disable=import-outside-toplevel

    workbook = xlsxwriter.Workbook(path)
    header_format = workbook.add_format({"bold": True, "bg_color": "#D9EAF7", "border": 0})
    integer_format = workbook.add_format({"num_format": "0"})
    decimal_format = workbook.add_format({"num_format": "0.000"})
    percent_format = workbook.add_format({"num_format": "0.00"})

    sections = (
        ("Wire Sizing Sweep", COARSE_HEADERS, result["coarse_rows"], "WireSizingSweep"),
        ("Transitions", TRANSITION_HEADERS, result["transition_rows"], "WireSizingTransitions"),
        ("Summary", SUMMARY_HEADERS, result["summary_rows"], "WireSizingSummary"),
    )
    for sheet_name, headers, rows, table_name in sections:
        worksheet = workbook.add_worksheet(_safe_sheet_name(sheet_name))
        worksheet.freeze_panes(1, 0)
        worksheet.write_row(0, 0, headers, header_format)
        for row_index, row in enumerate(rows, 1):
            values = _strip_private(row, headers)
            worksheet.write_row(row_index, 0, values)
        if rows:
            worksheet.add_table(
                0,
                0,
                len(rows),
                len(headers) - 1,
                {
                    "name": table_name,
                    "style": "Table Style Medium 2",
                    "columns": [{"header": header} for header in headers],
                },
            )
        worksheet.autofilter(0, 0, max(len(rows), 1), len(headers) - 1)
        worksheet.set_column(0, len(headers) - 1, 14)
        for column_index, header in enumerate(headers):
            if header in ("Status / Notes", "Configuration"):
                worksheet.set_column(column_index, column_index, 38)
            elif header in ("Hot Wire Size", "Ground Wire Size", "Baseline Hot Wire Size"):
                worksheet.set_column(column_index, column_index, 20)
            elif header in ("Wire Material", "Wire Temperature", "Test Point Type"):
                worksheet.set_column(column_index, column_index, 18)
            if header in ("Breaker Rating", "Load", "Length", "Poles", "Voltage", "Load Percent", "Hot Upsize Steps"):
                worksheet.set_column(column_index, column_index, 14, integer_format)
            if header in ("Base Ampacity",):
                worksheet.set_column(column_index, column_index, 14, decimal_format)
            if header in ("Voltage Drop %",):
                worksheet.set_column(column_index, column_index, 14, percent_format)
    workbook.close()


def _write_xlsx_openpyxl(path, result):
    from openpyxl import Workbook  # pylint: disable=import-outside-toplevel
    from openpyxl.styles import Font, PatternFill  # pylint: disable=import-outside-toplevel
    from openpyxl.worksheet.table import Table, TableStyleInfo  # pylint: disable=import-outside-toplevel

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    sections = (
        ("Wire Sizing Sweep", COARSE_HEADERS, result["coarse_rows"], "WireSizingSweep"),
        ("Transitions", TRANSITION_HEADERS, result["transition_rows"], "WireSizingTransitions"),
        ("Summary", SUMMARY_HEADERS, result["summary_rows"], "WireSizingSummary"),
    )
    for sheet_name, headers, rows, table_name in sections:
        worksheet = workbook.create_sheet(_safe_sheet_name(sheet_name))
        worksheet.freeze_panes = "A2"
        worksheet.append(list(headers))
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for row in rows:
            worksheet.append(_strip_private(row, headers))
        if rows:
            last_row = len(rows) + 1
            last_column = len(headers)
            ref = "A1:{}{}".format(
                openpyxl_column_name(last_column),
                last_row,
            )
            table = Table(displayName=table_name, ref=ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(table)
        for column_index, header in enumerate(headers, 1):
            width = 14
            if header in ("Status / Notes", "Configuration"):
                width = 38
            elif header in ("Hot Wire Size", "Ground Wire Size", "Baseline Hot Wire Size"):
                width = 20
            elif header in ("Wire Material", "Wire Temperature", "Test Point Type"):
                width = 18
            worksheet.column_dimensions[openpyxl_column_name(column_index)].width = width
    workbook.save(path)


def openpyxl_column_name(number):
    """Small dependency-free Excel column-name helper for the optional writer."""

    name = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        name = chr(65 + remainder) + name
    return name


def write_report(result, requested_path):
    """Write XLSX when available, otherwise a three-file CSV bundle."""

    requested_path = os.path.abspath(requested_path)
    output_dir = os.path.dirname(requested_path)
    if output_dir and not os.path.isdir(output_dir):
        os.makedirs(output_dir)

    extension = os.path.splitext(requested_path)[1].lower()
    if extension == ".xlsx":
        try:
            importlib.import_module("xlsxwriter")
            _write_xlsx_xlsxwriter(requested_path, result)
            return {"format": "xlsx", "primary_path": requested_path, "paths": [requested_path]}
        except ImportError:
            try:
                importlib.import_module("openpyxl")
                _write_xlsx_openpyxl(requested_path, result)
                return {"format": "xlsx", "primary_path": requested_path, "paths": [requested_path]}
            except ImportError:
                pass

    base_path = os.path.splitext(requested_path)[0]
    paths = [
        "{} - Wire Sizing Sweep.csv".format(base_path),
        "{} - Transitions.csv".format(base_path),
        "{} - Summary.csv".format(base_path),
    ]
    _write_csv(paths[0], COARSE_HEADERS, result["coarse_rows"])
    _write_csv(paths[1], TRANSITION_HEADERS, result["transition_rows"])
    _write_csv(paths[2], SUMMARY_HEADERS, result["summary_rows"])
    return {"format": "csv", "primary_path": paths[0], "paths": paths}


def print_summary(result, report):
    summary_rows = result["summary_rows"]
    all_rows = list(_iter_data_rows(result))
    failures = [row for row in all_rows if row.get("_status") in ("CALC_FAILED", "EXCEPTION")]
    five_upsizes = [row for row in summary_rows if row.get("Five Upsizes Reached") == "Yes"]
    upper_limits = [
        row
        for row in summary_rows
        if row.get("Adaptive Search Status") in ("UPPER_LIMIT_FIRST", "SAFETY_GUARD_REACHED")
    ]
    print("")
    print("Wire sizing sweep complete")
    print("Total rows generated: {}".format(len(all_rows)))
    print("Breaker ratings tested: {}".format(len(result["production"]["breakers"])))
    print("Configuration count: {}".format(len(ELECTRICAL_CONFIGURATIONS)))
    print("Failures: {}".format(len(failures)))
    print("Combinations reaching five upsizes: {}".format(len(five_upsizes)))
    print("Combinations hitting an upper limit first: {}".format(len(upper_limits)))
    print("Output format: {}".format(report["format"].upper()))
    print("Output file path: {}".format(report["primary_path"]))
    validation_failures = result.get("validation_failures") or []
    print("Structural validation failures: {}".format(len(validation_failures)))
    if validation_failures:
        display_limit = 12
        for failure in validation_failures[:display_limit]:
            print("  - {}".format(failure))
        if len(validation_failures) > display_limit:
            print("  ... and {} more; see validation_failures in the IDE result.".format(
                len(validation_failures) - display_limit
            ))


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT,
        help="Requested .xlsx path; falls back to three CSV files if no XLSX writer is installed.",
    )
    parser.add_argument(
        "--no-progress",
        action="store_true",
        help="Suppress per-combination progress messages.",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Raise an assertion after writing the report if structural validation fails.",
    )
    args = parser.parse_args(argv)

    result = generate_sweep(progress=not args.no_progress)
    report = write_report(result, args.output)
    print_summary(result, report)
    if args.strict:
        assert_result_integrity(result)
    return 0


if __name__ == "__main__":
    main()
